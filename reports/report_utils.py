# reports/report_utils.py
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import io
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime


def export_to_excel(data, title, sheet_name, headers, filename):
    """Export data to Excel with professional formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Define styles
    title_font = Font(name='Arial', size=16, bold=True, color='1a2a4a')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1e3c72', end_color='1e3c72', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # Add title
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    # Add generation date
    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    date_cell = ws['A2']
    date_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y %H:%M:%S')}"
    date_cell.font = Font(size=9, italic=True, color='666666')
    date_cell.alignment = center_alignment

    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Add data
    for row_idx, row_data in enumerate(data, start=5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            # Set alignment based on column type
            if col_idx == 1:
                cell.alignment = left_alignment
            elif isinstance(value, (int, float)) or (isinstance(value, str) and value.startswith('₱')):
                cell.alignment = right_alignment
            else:
                cell.alignment = center_alignment

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 18

    # Add footer note
    footer_row = len(data) + 6
    ws.merge_cells(f'A{footer_row}:{get_column_letter(len(headers))}{footer_row}')
    footer_cell = ws[f'A{footer_row}']
    footer_cell.value = "*** This is a computer-generated document. No signature is required. ***"
    footer_cell.font = Font(size=9, italic=True, color='888888')
    footer_cell.alignment = center_alignment

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


def export_to_pdf(data, title, headers, filename):
    """Export data to PDF with professional formatting - Center aligned header"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    # Create centered paragraph style
    center_style = ParagraphStyle(
        'CenterStyle',
        parent=styles['Normal'],
        alignment=1,  # 1 = CENTER
        fontSize=10,
        leading=14
    )

    # Title style - Centered
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a2a4a'),
        alignment=1,  # Center alignment
        spaceAfter=10,
        leading=20
    )

    # Subtitle style - Centered
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=13,
        textColor=colors.HexColor('#2a5298'),
        alignment=1,  # Center alignment
        spaceAfter=5,
        leading=18
    )

    # Add centered cooperative header
    cooperative_name = Paragraph('<font size="14" color="#1a2a4a"><b>TOLONG MULTI-PURPOSE COOPERATIVE</b></font>',
                                 title_style)
    cooperative_address = Paragraph(
        '<font size="9" color="#555555">National Highway, Villareal, Bayawan City, Negros Oriental</font>',
        center_style)
    cooperative_reg = Paragraph(
        '<font size="8" color="#777777">CDA Reg. No. 9520-0700431 | Established: November 13, 1992</font>',
        center_style)

    elements.append(cooperative_name)
    elements.append(Spacer(1, 0.05 * inch))
    elements.append(cooperative_address)
    elements.append(Spacer(1, 0.03 * inch))
    elements.append(cooperative_reg)
    elements.append(Spacer(1, 0.1 * inch))

    # Add title
    title_para = Paragraph(title, title_style)
    elements.append(title_para)
    elements.append(Spacer(1, 0.05 * inch))

    # Add generation date - Centered
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        alignment=1,  # Center alignment
        spaceAfter=15
    )
    date_text = f'Generated on: {datetime.now().strftime("%B %d, %Y %H:%M:%S")}'
    elements.append(Paragraph(date_text, date_style))
    elements.append(Spacer(1, 0.15 * inch))

    # Create table from data
    table_data = [headers]
    for row in data:
        table_data.append(row)

    # Calculate column widths
    col_widths = [1.5 * inch] + [1.2 * inch] * (len(headers) - 1)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3c72')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))

    # Add notes - Left aligned (for readability)
    notes_style = ParagraphStyle(
        'NotesStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#666666'),
        alignment=0,  # Left alignment
        leftIndent=10
    )
    notes = """
    <b>Notes to the Financial Statement:</b><br/>
    1. Interest is computed using the diminishing balance method at the applicable loan product rate.<br/>
    2. Penalty is applied at 2% per month after 360 days from the disbursement date.<br/>
    3. This report is generated automatically by the ToMPuCo Loan Management System.<br/>
    4. For inquiries or clarifications, please contact the cooperative office.
    """
    elements.append(Paragraph(notes, notes_style))
    elements.append(Spacer(1, 0.2 * inch))

    # Add signature lines - Centered
    signature_style = ParagraphStyle(
        'SignatureStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=1,  # Center alignment
        textColor=colors.HexColor('#333333')
    )

    # Signature section with three columns
    signature_data = [
        ['', '', ''],
        [
            'Prepared by:',
            'Noted by:',
            'Received by:'
        ],
        [
            f'{datetime.now().strftime("%Y-%m-%d")}',
            '_________________',
            '_________________'
        ],
        [
            f'{request.user.get_full_name() if hasattr(request, 'user') else "Staff Name"}',
            'Cooperative Manager',
            'Member/Representative'
        ]
    ]

    signature_table = Table(signature_data, colWidths=[2 * inch, 2 * inch, 2 * inch])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(signature_table)
    elements.append(Spacer(1, 0.15 * inch))

    # Add footer - Centered
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#888888'),
        alignment=1,  # Center alignment
        spaceBefore=10
    )
    footer_text = """
    This is a computer-generated document. No signature is required for validation.<br/>
    ToMPuCo Cooperative | Villareal, Bayawan City, Negros Oriental | Tel: (035) 123-4567<br/>
    *** This document is for official use only ***
    """
    elements.append(Paragraph(footer_text, footer_style))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    response.write(buffer.getvalue())
    return response