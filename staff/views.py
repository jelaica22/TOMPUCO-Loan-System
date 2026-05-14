from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt  # ← ADD THIS IMPORT
from datetime import datetime, timedelta, date
import json
from decimal import Decimal
import random
import secrets
import io
import base64
import qrcode

# 2FA imports
from django_otp.plugins.otp_totp.models import TOTPDevice

# Main models imports
from main.models import (
    Member, LoanProduct, LoanApplication, Loan, Payment,
    PaymentSchedule, MemberDocument, Notification, AuditLog
)

# Staff models
from .models import StaffProfile, StaffActivityLog, PaymentInstruction, RestructuringRequest, StaffNotification


# ==================== HELPER FUNCTIONS ====================

def is_staff(user):
    """Check if user has staff role"""
    if not user.is_authenticated:
        return False
    try:
        return hasattr(user, 'staff_profile') and user.staff_profile is not None
    except:
        return False


def staff_required(view_func):
    """Decorator to check if user is staff"""

    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('main:login')
        if not (request.user.is_staff or request.user.groups.filter(name='Staff').exists()):
            messages.error(request, 'Access denied. Staff privileges required.')
            return redirect('main:dashboard')
        return view_func(request, *args, **kwargs)

    return wrapper


def log_activity(staff, action, entity_type, entity_id, request):
    """Helper to log staff activities"""
    if StaffActivityLog:
        StaffActivityLog.objects.create(
            staff=staff,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )


def parse_report_dates(period, date_param):
    """Helper function to parse report period and date parameters"""
    today = date.today()

    if period == 'daily':
        if date_param:
            start_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        else:
            start_date = today
        end_date = start_date

    elif period == 'weekly':
        if date_param and '|' in date_param:
            dates = date_param.split('|')
            start_date = datetime.strptime(dates[0], '%Y-%m-%d').date()
            end_date = datetime.strptime(dates[1], '%Y-%m-%d').date()
        else:
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)

    elif period == 'monthly':
        if date_param:
            year, month = map(int, date_param.split('-'))
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
        else:
            start_date = date(today.year, today.month, 1)
            if today.month == 12:
                end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)

    elif period == 'quarterly':
        if date_param and 'Q' in date_param:
            year, quarter = date_param.split('-Q')
            year = int(year)
            quarter = int(quarter)
            start_month = (quarter - 1) * 3 + 1
            start_date = date(year, start_month, 1)
            if start_month + 2 > 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, start_month + 3, 1) - timedelta(days=1)
        else:
            quarter = (today.month - 1) // 3 + 1
            start_month = (quarter - 1) * 3 + 1
            start_date = date(today.year, start_month, 1)
            if start_month + 2 > 12:
                end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(today.year, start_month + 3, 1) - timedelta(days=1)

    elif period == 'yearly':
        if date_param:
            year = int(date_param)
            start_date = date(year, 1, 1)
            end_date = date(year, 12, 31)
        else:
            start_date = date(today.year, 1, 1)
            end_date = date(today.year, 12, 31)

    elif period == 'custom':
        if date_param and '|' in date_param:
            dates = date_param.split('|')
            start_date = datetime.strptime(dates[0], '%Y-%m-%d').date()
            end_date = datetime.strptime(dates[1], '%Y-%m-%d').date()
        else:
            start_date = today - timedelta(days=30)
            end_date = today
    else:
        start_date = today - timedelta(days=30)
        end_date = today

    return start_date, end_date


def generate_payment_schedule(loan):
    """Generate payment schedule for a loan"""
    from decimal import Decimal
    from datetime import timedelta

    monthly_payment = loan.monthly_payment
    remaining_balance = loan.amount
    due_date = loan.disbursement_date
    interest_rate = loan.interest_rate
    term_months = loan.term_months
    monthly_rate = interest_rate / Decimal('100') / Decimal('12')

    for month in range(1, term_months + 1):
        interest = remaining_balance * monthly_rate
        principal_paid = monthly_payment - interest

        if month == term_months:
            principal_paid = remaining_balance
            monthly_payment = principal_paid + interest

        PaymentSchedule.objects.create(
            loan=loan,
            schedule_number=month,
            due_date=due_date,
            principal_due=principal_paid,
            interest_due=interest,
            total_due=monthly_payment,
            status='pending'
        )

        remaining_balance -= principal_paid
        due_date = due_date + timedelta(days=30)


# ==================== DASHBOARD ====================

@login_required
@staff_required
def dashboard(request):
    """Staff dashboard with analytics"""
    from datetime import date, timedelta
    from django.db.models import Sum, Count, Q

    staff_profile = request.user.staff_profile
    today = date.today()

    # ============================================================
    # STATS CARDS
    # ============================================================
    total_applications = LoanApplication.objects.count()
    pending_count = LoanApplication.objects.filter(status='pending_staff_review').count()
    active_loans = Loan.objects.filter(status='active').count()
    overdue_loans = Loan.objects.filter(status='overdue').count()

    # ============================================================
    # FINANCIAL SUMMARY
    # ============================================================
    total_principal = Loan.objects.aggregate(total=Sum('amount'))['total'] or 0
    total_paid = Payment.objects.filter(status='completed').aggregate(total=Sum('amount'))['total'] or 0
    collection_rate = (float(total_paid) / float(total_principal) * 100) if total_principal > 0 else 0

    # Today's collection
    today_collection = Payment.objects.filter(
        payment_date=today,
        status='completed'
    ).aggregate(total=Sum('amount'))['total'] or 0

    # ============================================================
    # CHART DATA
    # ============================================================
    trend_labels = []
    trend_data = []
    for i in range(6, -1, -1):
        date_val = today - timedelta(days=i)
        trend_labels.append(date_val.strftime('%a, %b %d'))
        count = LoanApplication.objects.filter(applied_date=date_val).count()
        trend_data.append(count)

    distribution_labels = ['Active', 'Overdue', 'Paid', 'Restructured']
    distribution_data = [
        Loan.objects.filter(status='active').count(),
        Loan.objects.filter(status='overdue').count(),
        Loan.objects.filter(status='paid').count(),
        Loan.objects.filter(status='restructured').count(),
    ]

    # ============================================================
    # RECENT ITEMS
    # ============================================================
    recent_applications = LoanApplication.objects.select_related('member', 'loan_product').order_by('-applied_date')[:5]
    recent_loans = Loan.objects.select_related('member').order_by('-created_at')[:5]

    # ============================================================
    # QUICK STATS
    # ============================================================
    month_start = date(today.year, today.month, 1)
    monthly_applications = LoanApplication.objects.filter(applied_date__gte=month_start).count()
    approved_count = LoanApplication.objects.filter(status='manager_approved').count()
    approval_rate = int((approved_count / total_applications * 100)) if total_applications > 0 else 0
    avg_loan_size = total_principal / total_applications if total_applications > 0 else 0
    disbursed_today = Loan.objects.filter(disbursement_date=today).aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'staff_profile': staff_profile,
        'total_applications': total_applications,
        'pending_count': pending_count,
        'active_loans': active_loans,
        'overdue_loans': overdue_loans,
        'total_principal': total_principal,
        'total_paid': total_paid,
        'collection_rate': round(collection_rate, 2),
        'today_collection': today_collection,
        'trend_labels': trend_labels,
        'trend_data': trend_data,
        'distribution_labels': distribution_labels,
        'distribution_data': distribution_data,
        'recent_applications': recent_applications,
        'recent_loans': recent_loans,
        'monthly_applications': monthly_applications,
        'approval_rate': approval_rate,
        'avg_loan_size': avg_loan_size,
        'disbursed_today': disbursed_today,
    }

    return render(request, 'staff/dashboard.html', context)


# ==================== APPLICATION VIEWS ====================

@login_required
@staff_required
def applications_list(request):
    """List all loan applications for staff"""
    staff_profile = request.user.staff_profile

    # Get all applications with related data
    applications = LoanApplication.objects.select_related('member', 'loan_product').order_by('-applied_date')

    # Debug - print to console to verify
    print(f"Total applications: {applications.count()}")
    for app in applications:
        print(f"App {app.id}: {app.application_id} - Amount: {app.amount} - Date: {app.applied_date}")

    context = {
        'staff_profile': staff_profile,
        'applications': applications,
        'total_applications': applications.count(),
        'pending_count': applications.filter(status='pending_staff_review').count(),
        'committee_count': applications.filter(status='with_committee').count(),
        'approved_count': applications.filter(status='manager_approved').count(),
        'rejected_count': applications.filter(status='rejected').count(),
    }
    return render(request, 'staff/applications/list.html', context)


@login_required
@staff_required
def create_application(request):
    """Create a new loan application"""
    staff_profile = request.user.staff_profile

    if request.method == 'POST':
        try:
            member_id = request.POST.get('member_id')
            loan_product_id = request.POST.get('loan_product_id')
            requested_amount = request.POST.get('requested_amount')
            purpose = request.POST.get('purpose')
            collateral = request.POST.get('collateral', '')
            payment_mode = request.POST.get('payment_mode', 'cash')
            notes = request.POST.get('notes', '')

            if not all([member_id, loan_product_id, requested_amount, purpose]):
                messages.error(request, 'Please fill in all required fields.')
                context = {
                    'staff_profile': staff_profile,
                    'loan_products': LoanProduct.objects.filter(is_active=True),
                    'members': Member.objects.filter(is_active=True)
                }
                return render(request, 'staff/applications/create.html', context)

            member = Member.objects.get(id=member_id)
            loan_product = LoanProduct.objects.get(id=loan_product_id)

            app_id = f"{loan_product.name[:4]}-{datetime.now().year}-{random.randint(1000, 9999)}"

            application = LoanApplication.objects.create(
                application_id=app_id,
                member=member,
                loan_product=loan_product,
                amount=Decimal(requested_amount),
                purpose=purpose,
                collateral=collateral,
                payment_mode=payment_mode,
                notes=notes,
                status='pending_staff_review',
                applied_date=timezone.now().date(),
                applicant_user=request.user
            )

            log_activity(staff_profile, 'CREATE', 'application', application.id, request)
            messages.success(request, f'Application {application.application_id} created successfully!')
            return redirect('staff:review_application', pk=application.id)

        except Member.DoesNotExist:
            messages.error(request, 'Selected member does not exist.')
        except LoanProduct.DoesNotExist:
            messages.error(request, 'Selected loan product does not exist.')
        except Exception as e:
            messages.error(request, f'Error creating application: {str(e)}')

        return redirect('staff:staff_applications')

    context = {
        'staff_profile': staff_profile,
        'loan_products': LoanProduct.objects.filter(is_active=True),
        'members': Member.objects.filter(is_active=True),
    }
    return render(request, 'staff/applications/create.html', context)


@login_required
@staff_required
def application_review(request, pk):
    """Review a specific application"""
    staff_profile = request.user.staff_profile
    application = get_object_or_404(LoanApplication, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'forward_to_committee':
            application.status = 'with_committee'
            application.remarks = request.POST.get('remarks', '')
            application.save()
            messages.success(request, f'Application forwarded to Committee')
        elif action == 'request_revision':
            application.status = 'needs_revision'
            application.remarks = request.POST.get('revision_reason', '')
            application.save()
            messages.info(request, f'Revision requested')
        elif action == 'reject':
            application.status = 'rejected'
            application.remarks = request.POST.get('reject_reason', '')
            application.save()
            messages.warning(request, f'Application rejected')

        return redirect('staff:staff_applications')

    context = {
        'staff_profile': staff_profile,
        'application': application,
    }
    return render(request, 'staff/applications/review.html', context)


@login_required
@staff_required
def edit_application(request, pk):
    """Edit application"""
    staff_profile = request.user.staff_profile
    application = get_object_or_404(LoanApplication, pk=pk)

    if request.method == 'POST':
        application.loan_product_id = request.POST.get('loan_product_id', application.loan_product_id)
        application.amount = request.POST.get('amount', application.amount)
        application.purpose = request.POST.get('purpose', application.purpose)
        application.collateral = request.POST.get('collateral', application.collateral)
        application.save()

        messages.success(request, f'Application {application.application_id} updated successfully!')
        return redirect('staff:staff_applications')

    loan_products = LoanProduct.objects.filter(is_active=True)

    context = {
        'staff_profile': staff_profile,
        'application': application,
        'loan_products': loan_products,
    }
    return render(request, 'staff/applications/edit.html', context)


@login_required
@staff_required
def add_charges(request, pk):
    """Add charges based on approved line (after manager approval)"""
    staff_profile = request.user.staff_profile
    application = get_object_or_404(LoanApplication, pk=pk)

    if application.status != 'manager_approved':
        messages.error(request, 'Application must be approved by manager before adding charges.')
        return redirect('staff:staff_applications')

    approved_line = float(application.approved_line or 0)

    if request.method == 'POST':
        service_charge = approved_line * 0.03
        cbu_retention = approved_line * 0.02
        insurance = approved_line * 0.0132
        service_fee = 35
        notarial_fee = 200

        inspector_fee = float(request.POST.get('inspector_fee', 0))
        trade_fert = float(request.POST.get('trade_fert', 0))
        ca_int = float(request.POST.get('ca_int', 0))

        total_auto = service_charge + cbu_retention + insurance + service_fee + notarial_fee
        total_optional = inspector_fee + trade_fert + ca_int
        total_deductions = total_auto + total_optional
        net_proceeds = approved_line - total_deductions  # This calculates correctly

        # Save all values
        application.service_charge = service_charge
        application.cbu_retention = cbu_retention
        application.insurance_charge = insurance
        application.service_fee = service_fee
        application.notarial_fee = notarial_fee
        application.inspector_fee = inspector_fee
        application.trade_fert = trade_fert
        application.ca_int = ca_int
        application.total_deductions = total_deductions
        application.net_proceeds = net_proceeds  # This should be saved
        application.status = 'ready_for_disbursement'
        application.save()

        messages.success(request, f'Charges added. Net Proceeds: ₱{net_proceeds:,.2f}')
        return redirect('staff:staff_applications')

    # GET request - show form
    service_charge = approved_line * 0.03
    cbu_retention = approved_line * 0.02
    insurance = approved_line * 0.0132
    service_fee = 35
    notarial_fee = 200
    total_auto = service_charge + cbu_retention + insurance + service_fee + notarial_fee
    net_proceeds = approved_line - total_auto

    context = {
        'staff_profile': staff_profile,
        'application': application,
        'approved_line': approved_line,
        'service_charge': service_charge,
        'cbu_retention': cbu_retention,
        'insurance': insurance,
        'service_fee': service_fee,
        'notarial_fee': notarial_fee,
        'total_auto': total_auto,
        'net_proceeds': net_proceeds,
    }

    return render(request, 'staff/add_charges.html', context)


@login_required
@staff_required
def create_loan(request, app_id):
    """Create loan from approved application"""
    application = get_object_or_404(LoanApplication, id=app_id)

    if application.status != 'manager_approved':
        messages.error(request, 'This application is not approved yet.')
        return redirect('staff:applications_list')

    existing_loan = Loan.objects.filter(application=application).first()
    if existing_loan:
        messages.warning(request, f'Loan already created: {existing_loan.loan_number}')
        return redirect('staff:loan_detail', loan_id=existing_loan.id)

    if request.method == 'POST':
        try:
            principal = application.approved_line if application.approved_line else application.amount
            term_months = int(request.POST.get('term_months', 12))

            loan_product = application.loan_product
            interest_rate = Decimal(str(loan_product.interest_rate)) if loan_product else Decimal('20.00')

            monthly_rate = interest_rate / Decimal('100') / Decimal('12')
            if monthly_rate > 0:
                monthly_payment = principal * (monthly_rate * (1 + monthly_rate) ** term_months) / (
                            (1 + monthly_rate) ** term_months - 1)
            else:
                monthly_payment = principal / term_months

            loan_number = f"LN-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"

            loan = Loan.objects.create(
                member=application.member,
                application=application,
                loan_number=loan_number,
                amount=principal,
                interest_rate=interest_rate,
                term_months=term_months,
                monthly_payment=monthly_payment,
                remaining_balance=principal,
                status='active',
                disbursement_date=timezone.now().date(),
                due_date=timezone.now().date() + timedelta(days=term_months * 30),
                created_at=timezone.now()
            )

            generate_payment_schedule(loan)

            application.status = 'disbursed'
            application.save()

            Notification.objects.create(
                recipient=application.member.user,
                notification_type='loan',
                title='🎉 Loan Disbursed Successfully!',
                message=f'Your loan of ₱{principal:,.2f} has been disbursed.',
                link=f'/member/loans/{loan.id}/'
            )

            messages.success(request, f'✅ Loan #{loan.loan_number} created successfully!')
            return redirect('staff:loan_detail', loan_id=loan.id)

        except Exception as e:
            messages.error(request, f'Error creating loan: {str(e)}')
            return redirect('staff:applications_list')

    principal = application.approved_line if application.approved_line else application.amount
    loan_product = application.loan_product
    interest_rate = Decimal(str(loan_product.interest_rate)) if loan_product else Decimal('20.00')

    service_charge = principal * Decimal('0.03')
    cbu_retention = principal * Decimal('0.02')
    insurance = principal * Decimal('0.0132')
    service_fee = Decimal('35.00')
    notarial_fee = Decimal('200.00')

    total_deductions = service_charge + cbu_retention + insurance + service_fee + notarial_fee
    net_proceeds = principal - total_deductions

    context = {
        'application': application,
        'principal': float(principal),
        'interest_rate': float(interest_rate),
        'net_proceeds': float(net_proceeds),
        'total_deductions': float(total_deductions),
        'service_charge': float(service_charge),
        'cbu_retention': float(cbu_retention),
        'insurance': float(insurance),
        'service_fee': float(service_fee),
        'notarial_fee': float(notarial_fee),
    }

    return render(request, 'staff/create_loan.html', context)


# ==================== LOAN VIEWS ====================

@login_required
@staff_required
def loan_list(request):
    """Loan register - list all loans"""
    staff_profile = request.user.staff_profile

    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'all')
    loan_type = request.GET.get('loan_type', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    loans = Loan.objects.select_related('member').all()

    if search:
        loans = loans.filter(
            Q(loan_number__icontains=search) |
            Q(member__first_name__icontains=search) |
            Q(member__last_name__icontains=search)
        )

    if status_filter != 'all':
        loans = loans.filter(status=status_filter)

    if loan_type != 'all':
        loans = loans.filter(loan_product__name=loan_type)

    if date_from:
        loans = loans.filter(disbursement_date__gte=date_from)
    if date_to:
        loans = loans.filter(disbursement_date__lte=date_to)

    today = date.today()
    for loan in loans:
        if loan.due_date and loan.due_date < today:
            loan.days_overdue = (today - loan.due_date).days
            if loan.days_overdue > 360:
                penalty_months = ((loan.days_overdue - 360) + 29) // 30
                loan.penalty_amount = float(loan.remaining_balance or 0) * 0.02 * penalty_months
            else:
                loan.penalty_amount = 0
        else:
            loan.days_overdue = 0
            loan.penalty_amount = 0

    context = {
        'staff_profile': staff_profile,
        'loans': loans,
        'active_loans_count': loans.filter(status='active').count(),
        'overdue_loans_count': loans.filter(status='overdue').count(),
        'restructured_loans_count': loans.filter(status='restructured').count(),
        'paid_loans_count': loans.filter(status='paid').count(),
        'total_loans': loans.count(),
        'total_principal': loans.aggregate(total=Sum('amount'))['total'] or 0,
    }

    return render(request, 'staff/loans/list.html', context)


@login_required
@staff_required
def loan_detail(request, pk):
    """Loan details with payment schedule"""
    staff_profile = request.user.staff_profile
    loan = get_object_or_404(Loan, pk=pk)

    context = {
        'staff_profile': staff_profile,
        'loan': loan,
    }

    return render(request, 'staff/loans/detail.html', context)


@login_required
@staff_required
def loan_payment_schedule(request, pk):
    """View loan payment schedule"""
    loan = get_object_or_404(Loan, pk=pk)

    schedule = []
    balance = float(loan.remaining_balance)
    monthly_payment = float(loan.monthly_payment)
    interest_rate = float(loan.interest_rate)
    term_months = loan.term_months
    monthly_rate = interest_rate / 100 / term_months

    next_date = loan.due_date or (loan.disbursement_date + timedelta(days=30))

    for i in range(term_months):
        interest = balance * monthly_rate
        principal = monthly_payment - interest
        if principal > balance:
            principal = balance
            monthly_payment = interest + principal

        balance = balance - principal

        schedule.append({
            'month': i + 1,
            'due_date': next_date.strftime('%Y-%m-%d'),
            'payment_amount': round(monthly_payment, 2),
            'interest': round(interest, 2),
            'principal': round(principal, 2),
            'balance': round(max(0, balance), 2)
        })

        next_date = next_date + timedelta(days=30)
        if balance <= 0:
            break

    return JsonResponse({'schedule': schedule})


@login_required
@staff_required
def process_restructuring(request, pk):
    """Process loan restructuring"""
    loan = get_object_or_404(Loan, pk=pk)

    if request.method == 'POST':
        new_principal = request.POST.get('new_principal')
        new_term = request.POST.get('new_term', 12)
        reason = request.POST.get('reason')

        old_balance = float(loan.remaining_balance)
        penalty = float(loan.penalty_amount or 0)

        restructure = RestructuringRequest.objects.create(
            request_number=f"RST-{datetime.now().strftime('%Y%m%d')}-{RestructuringRequest.objects.count() + 1:04d}",
            member_id=loan.member.id,
            old_loan_id=loan.id,
            old_balance=old_balance,
            old_penalty=penalty,
            new_principal=new_principal,
            new_interest_rate=loan.interest_rate,
            new_term_months=new_term,
            reason=reason,
            status='pending_staff',
            staff_approved_by=request.user.staff_profile
        )

        messages.success(request, f'Restructuring request {restructure.request_number} created!')
        return redirect('staff:restructuring_list')

    return JsonResponse({'error': 'Invalid method'}, status=400)


@login_required
@staff_required
def export_loans(request):
    """Export loans to Excel"""
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Register"

    headers = ['Loan #', 'Borrower', 'Loan Type', 'Principal', 'Interest Rate', 'Monthly Payment',
               'Remaining Balance', 'Status', 'Disbursement Date', 'Term (Months)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1e3c72', end_color='1e3c72', fill_type='solid')

    loans = Loan.objects.select_related('member').all()
    for row, loan in enumerate(loans, 2):
        ws.cell(row=row, column=1, value=loan.loan_number)
        ws.cell(row=row, column=2, value=f"{loan.member.last_name}, {loan.member.first_name}")
        ws.cell(row=row, column=3, value=loan.loan_product.name if loan.loan_product else 'N/A')
        ws.cell(row=row, column=4, value=float(loan.amount))
        ws.cell(row=row, column=5, value=float(loan.interest_rate))
        ws.cell(row=row, column=6, value=float(loan.monthly_payment))
        ws.cell(row=row, column=7, value=float(loan.remaining_balance))
        ws.cell(row=row, column=8, value=loan.status)
        ws.cell(row=row, column=9, value=loan.disbursement_date.strftime('%Y-%m-%d') if loan.disbursement_date else '')
        ws.cell(row=row, column=10, value=loan.term_months)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="loans_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)

    return response


# ==================== PAYMENT VIEWS ====================

@login_required
@staff_required
def payment_list(request):
    """Payments page"""
    staff_profile = request.user.staff_profile

    context = {
        'staff_profile': staff_profile,
        'today_collection': 0,
        'week_collection': 0,
        'month_collection': 0,
        'pending_instructions': 0,
        'recent_payments': [],
    }

    return render(request, 'staff/payments/list.html', context)


@login_required
@staff_required
def payment_history(request):
    """Payment history page"""
    staff_profile = request.user.staff_profile
    payments = Payment.objects.filter(is_posted=True).order_by('-payment_date')[:50]

    context = {
        'staff_profile': staff_profile,
        'payments': payments,
    }

    return render(request, 'staff/payments/decision_history.html', context)


@login_required
@staff_required
def payment_detail(request, pk):
    """Payment detail page - also handles JSON requests"""
    staff_profile = request.user.staff_profile
    payment = get_object_or_404(Payment, pk=pk)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return JsonResponse({
            'id': payment.id,
            'payment_number': payment.payment_number,
            'payment_date': payment.payment_date.strftime('%Y-%m-%d'),
            'member_name': f"{payment.member.last_name}, {payment.member.first_name}" if payment.member else "Unknown",
            'member_id': payment.member.membership_number if payment.member else "N/A",
            'loan_number': payment.loan.loan_number if payment.loan else "N/A",
            'payment_method': payment.payment_method,
            'amount': float(payment.amount),
            'penalty_amount': float(payment.penalty_amount) if hasattr(payment, 'penalty_amount') else 0,
            'interest_amount': float(payment.interest_amount) if hasattr(payment, 'interest_amount') else 0,
            'principal_amount': float(payment.principal_amount) if hasattr(payment, 'principal_amount') else 0,
            'remaining_balance': float(payment.remaining_balance_after) if hasattr(payment,
                                                                                   'remaining_balance_after') else 0,
            'status': payment.status,
            'posted_by': payment.posted_by.get_full_name() if payment.posted_by else "Cashier",
        })

    context = {
        'staff_profile': staff_profile,
        'payment': payment,
    }
    return render(request, 'staff/payments/detail.html', context)


@login_required
@staff_required
def issue_payment_instruction(request):
    """Issue payment instruction page"""
    staff_profile = request.user.staff_profile

    context = {
        'staff_profile': staff_profile,
    }

    return render(request, 'staff/payments/issue.html', context)


@csrf_exempt
@login_required
@staff_required
def issue_instruction_api(request):
    """API endpoint to issue payment instruction"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            instruction = PaymentInstruction.objects.create(
                instruction_number=f"PI-{datetime.now().strftime('%Y%m%d%H%M%S')}",
                member_id=data.get('member_id'),
                loan_id=data.get('loan_id'),
                amount_to_collect=data.get('amount'),
                penalty_amount=data.get('penalty', 0),
                interest_amount=data.get('interest', 0),
                principal_amount=data.get('principal', 0),
                remaining_balance_after=data.get('remaining_balance', 0),
                next_due_date=datetime.now().date() + timedelta(days=30),
                issued_by=request.user.staff_profile
            )

            return JsonResponse({
                'success': True,
                'instruction_number': instruction.instruction_number,
                'amount': float(instruction.amount_to_collect),
                'penalty': float(instruction.penalty_amount),
                'interest': float(instruction.interest_amount),
                'principal': float(instruction.principal_amount),
                'remaining_balance': float(instruction.remaining_balance_after)
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'error': 'Invalid method'}, status=400)


@login_required
@staff_required
def payment_receipt_json(request, pk):
    """Return payment receipt data as JSON"""
    payment = get_object_or_404(Payment, pk=pk)

    data = {
        'payment_number': payment.payment_number,
        'payment_date': payment.payment_date.strftime('%Y-%m-%d'),
        'member_name': f"{payment.member.last_name}, {payment.member.first_name}" if payment.member else "Unknown",
        'member_id': payment.member.membership_number if payment.member else "N/A",
        'loan_number': payment.loan.loan_number if payment.loan else "N/A",
        'payment_method': payment.payment_method,
        'amount': float(payment.amount),
        'penalty_amount': 0,
        'interest_amount': 0,
        'principal_amount': float(payment.amount),
        'remaining_balance': float(payment.remaining_balance_after) if hasattr(payment,
                                                                               'remaining_balance_after') else 0,
        'status': payment.status,
        'posted_by': payment.posted_by.get_full_name() if payment.posted_by else "Cashier",
    }

    return JsonResponse(data)


# ==================== RESTRUCTURING VIEWS ====================

@login_required
@staff_required
def restructuring_list(request):
    """Restructuring Management page"""
    staff_profile = request.user.staff_profile

    pending_count = RestructuringRequest.objects.filter(status__in=['pending_staff', 'with_committee']).count()
    approved_count = RestructuringRequest.objects.filter(status__in=['committee_approved', 'manager_approved']).count()
    completed_count = RestructuringRequest.objects.filter(status='completed').count()
    total_requests = RestructuringRequest.objects.count()

    context = {
        'staff_profile': staff_profile,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'completed_count': completed_count,
        'total_requests': total_requests,
    }

    return render(request, 'staff/restructuring/list.html', context)


@login_required
@staff_required
def restructuring_detail(request, pk):
    """Restructuring detail page"""
    staff_profile = request.user.staff_profile
    restructure = get_object_or_404(RestructuringRequest, pk=pk)

    context = {
        'staff_profile': staff_profile,
        'restructure': restructure,
    }

    return render(request, 'staff/restructuring/detail.html', context)


@login_required
@staff_required
def restructuring_request_form(request, member_id=None):
    """Restructuring request form"""
    staff_profile = request.user.staff_profile

    context = {
        'staff_profile': staff_profile,
        'member_id': member_id,
    }

    return render(request, 'staff/restructuring/request.html', context)


@csrf_exempt
@login_required
@staff_required
def restructuring_api_request(request):
    """API endpoint to submit a restructuring request"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            if not data.get('member_id'):
                return JsonResponse({'success': False, 'error': 'Member ID is required'})
            if not data.get('loan_id'):
                return JsonResponse({'success': False, 'error': 'Loan ID is required'})
            if not data.get('reason'):
                return JsonResponse({'success': False, 'error': 'Reason for restructuring is required'})

            calc = data.get('calculation', {})
            request_number = f"RST-{datetime.now().strftime('%Y%m%d')}-{str(datetime.now().timestamp())[-6:]}"

            restructuring = RestructuringRequest.objects.create(
                request_number=request_number,
                member_id=data.get('member_id'),
                old_loan_id=data.get('loan_id'),
                old_balance=calc.get('old_balance', 0),
                old_interest=calc.get('interest_due', 0),
                old_penalty=calc.get('penalty', 0),
                days_overdue=calc.get('days_overdue', 0),
                new_charges=calc.get('total_new_charges', 0),
                new_principal=calc.get('new_principal', 0),
                new_interest_rate=calc.get('interest_rate', 20),
                new_term_months=12,
                new_monthly_payment=calc.get('monthly_payment', 0),
                reason=data.get('reason', ''),
                status='pending_staff',
                staff_approved_by=request.user.staff_profile,
                staff_approved=True
            )

            return JsonResponse({
                'success': True,
                'request_number': request_number,
                'request_id': restructuring.id
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'error': 'Invalid method'}, status=400)


@login_required
@staff_required
def restructuring_api_list(request):
    """API endpoint to get list of restructuring requests"""
    requests = RestructuringRequest.objects.all().order_by('-created_at')[:50]

    result = []
    for req in requests:
        try:
            member = Member.objects.get(id=req.member_id)
            member_name = f"{member.last_name}, {member.first_name}"
        except:
            member_name = "Unknown"

        try:
            loan = Loan.objects.get(id=req.old_loan_id)
            loan_number = loan.loan_number
        except:
            loan_number = "N/A"

        result.append({
            'id': req.id,
            'request_number': req.request_number,
            'member_name': member_name,
            'loan_number': loan_number,
            'old_balance': float(req.old_balance),
            'new_principal': float(req.new_principal),
            'new_monthly_payment': float(req.new_monthly_payment),
            'status': req.status,
            'created_at': req.created_at.strftime('%Y-%m-%d %H:%M'),
        })

    return JsonResponse({'requests': result})


@login_required
@staff_required
def restructuring_api_detail(request, pk):
    """API endpoint to get details of a specific restructuring request"""
    restructure = get_object_or_404(RestructuringRequest, pk=pk)

    try:
        member = Member.objects.get(id=restructure.member_id)
        member_name = f"{member.last_name}, {member.first_name}"
        membership_number = member.membership_number
    except:
        member_name = "Unknown"
        membership_number = "N/A"

    try:
        loan = Loan.objects.get(id=restructure.old_loan_id)
        loan_number = loan.loan_number
        original_principal = float(loan.amount)
    except:
        loan_number = "N/A"
        original_principal = 0

    return JsonResponse({
        'id': restructure.id,
        'request_number': restructure.request_number,
        'member_name': member_name,
        'membership_number': membership_number,
        'loan_number': loan_number,
        'original_principal': original_principal,
        'old_balance': float(restructure.old_balance),
        'old_interest': float(restructure.old_interest),
        'old_penalty': float(restructure.old_penalty),
        'new_principal': float(restructure.new_principal),
        'new_interest_rate': float(restructure.new_interest_rate),
        'new_monthly_payment': float(restructure.new_monthly_payment),
        'reason': restructure.reason,
        'status': restructure.status,
        'created_at': restructure.created_at.strftime('%Y-%m-%d %H:%M'),
    })


# ==================== REPORT VIEWS ====================

@login_required
@staff_required
def reports_index(request):
    """Reports dashboard"""
    staff_profile = request.user.staff_profile

    context = {
        'staff_profile': staff_profile,
        'total_loans': Loan.objects.count(),
        'total_collection': 0,
        'total_members': Member.objects.count(),
        'overdue_loans': Loan.objects.filter(status='overdue').count(),
    }

    return render(request, 'staff/reports/index.html', context)


@login_required
@staff_required
def report_loan_summary(request):
    """API endpoint for loan summary report"""
    return JsonResponse({
        'loans': [],
        'total': {'count': 0, 'principal': 0, 'interest': 0, 'total': 0, 'collection_rate': 0}
    })


@login_required
@staff_required
def report_collection(request):
    """API endpoint for collection report"""
    return JsonResponse({'daily': [], 'total': {'count': 0, 'cash': 0, 'quedan': 0, 'pesada': 0, 'total': 0}})


@login_required
@staff_required
def report_aging(request):
    """API endpoint for aging report"""
    return JsonResponse(
        {'categories': [], 'total': {'count': 0, 'principal': 0, 'interest': 0, 'penalty': 0, 'total': 0}})


@login_required
@staff_required
def report_member(request):
    """API endpoint for member report"""
    return JsonResponse({'types': [], 'total': {'count': 0, 'loan_count': 0, 'borrowed': 0, 'paid': 0, 'remaining': 0}})


@login_required
@staff_required
def report_loan_product(request):
    """API endpoint for loan product report"""
    return JsonResponse({'products': []})


@login_required
@staff_required
def report_restructuring(request):
    """API endpoint for restructuring report"""
    return JsonResponse({'requests': []})


@login_required
@staff_required
def report_penalty(request):
    """API endpoint for penalty report"""
    return JsonResponse({'penalties': [], 'total_penalty': 0})


@login_required
@staff_required
def export_report_excel(request):
    """Export report to Excel"""
    report_type = request.GET.get('report_type', 'loan-summary')
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
    response.write(f"Report: {report_type}\nGenerated: {datetime.now()}")
    return response


@login_required
@staff_required
def export_report_pdf(request):
    """Export report to PDF"""
    report_type = request.GET.get('report_type', 'loan-summary')
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.pdf"'
    response.write(f"Report: {report_type}\nGenerated: {datetime.now()}")
    return response


# ==================== REPORT API ENDPOINTS ====================

@login_required
@staff_required
def report_loan_summary_api(request):
    """API endpoint for loan summary report"""
    period = request.GET.get('period', 'monthly')
    date_param = request.GET.get('date', '')
    start_date, end_date = parse_report_dates(period, date_param)

    loans = Loan.objects.filter(disbursement_date__gte=start_date, disbursement_date__lte=end_date)
    total_principal = loans.aggregate(total=Sum('amount'))['total'] or 0

    return JsonResponse({
        'loans': [],
        'total': {'count': loans.count(), 'principal': float(total_principal), 'interest': 0,
                  'total': float(total_principal), 'collection_rate': 0}
    })


@login_required
@staff_required
def report_collection_api(request):
    """API endpoint for collection report"""
    period = request.GET.get('period', 'monthly')
    date_param = request.GET.get('date', '')
    start_date, end_date = parse_report_dates(period, date_param)

    payments = Payment.objects.filter(payment_date__gte=start_date, payment_date__lte=end_date, status='completed')

    return JsonResponse({
        'daily': [],
        'total': {'count': payments.count(), 'cash': 0, 'quedan': 0, 'pesada': 0, 'total': 0}
    })


@login_required
@staff_required
def report_aging_api(request):
    """API endpoint for aging report"""
    return JsonResponse(
        {'categories': [], 'total': {'count': 0, 'principal': 0, 'interest': 0, 'penalty': 0, 'total': 0}})


@login_required
@staff_required
def report_member_api(request):
    """API endpoint for member report"""
    return JsonResponse({'types': [], 'total': {'count': 0, 'loan_count': 0, 'borrowed': 0, 'paid': 0, 'remaining': 0}})


@login_required
@staff_required
def report_loan_product_api(request):
    """API endpoint for loan product report"""
    return JsonResponse({'products': []})


@login_required
@staff_required
def report_restructuring_api(request):
    """API endpoint for restructuring report"""
    period = request.GET.get('period', 'monthly')
    date_param = request.GET.get('date', '')
    start_date, end_date = parse_report_dates(period, date_param)

    restructures = RestructuringRequest.objects.filter(created_at__gte=start_date, created_at__lte=end_date)
    request_data = []

    for req in restructures:
        try:
            member = Member.objects.get(id=req.member_id)
            member_name = f"{member.last_name}, {member.first_name}"
        except:
            member_name = "Unknown"

        request_data.append({
            'number': req.request_number,
            'member': member_name,
            'old_balance': float(req.old_balance),
            'new_principal': float(req.new_principal),
            'status': req.status,
            'date': req.created_at.strftime('%Y-%m-%d')
        })

    return JsonResponse({'requests': request_data})


@login_required
@staff_required
def report_penalty_api(request):
    """API endpoint for penalty report"""
    today = date.today()
    overdue_loans = Loan.objects.filter(status='overdue')
    penalties = []
    total_penalty = 0

    for loan in overdue_loans:
        days_overdue = 0
        if loan.due_date and loan.due_date < today:
            days_overdue = (today - loan.due_date).days

        if days_overdue > 360:
            penalty_months = ((days_overdue - 360) + 29) // 30
            penalty = float(loan.remaining_balance or 0) * 0.02 * penalty_months
            total_penalty += penalty
            penalties.append({
                'loan_number': loan.loan_number,
                'member': f"{loan.member.last_name}, {loan.member.first_name}",
                'balance': float(loan.remaining_balance or 0),
                'days_overdue': days_overdue,
                'total_penalty': penalty
            })

    return JsonResponse({'penalties': penalties, 'total_penalty': total_penalty})


# ==================== NOTIFICATION VIEWS ====================

@login_required
@staff_required
def notifications_page(request):
    """Notifications page"""
    staff_profile = request.user.staff_profile

    notifications = StaffNotification.objects.filter(staff=staff_profile).order_by('-created_at')

    context = {
        'staff_profile': staff_profile,
        'notifications': notifications[:20],
        'total_notifications': notifications.count(),
        'unread_count': notifications.filter(is_read=False).count(),
        'read_count': notifications.filter(is_read=True).count(),
        'system_notifications': notifications.filter(notification_type='system_alert').count(),
    }

    return render(request, 'staff/notifications/list.html', context)


@login_required
@staff_required
def notifications_api(request):
    """API endpoint for notifications"""
    staff_profile = request.user.staff_profile
    unread_count = StaffNotification.objects.filter(staff=staff_profile, is_read=False).count()
    return JsonResponse({'unread_count': unread_count})


@login_required
@staff_required
def mark_notification_read(request, pk):
    """Mark a single notification as read"""
    if request.method == 'POST':
        try:
            notification = StaffNotification.objects.get(pk=pk, staff=request.user.staff_profile)
            notification.is_read = True
            notification.read_at = timezone.now()
            notification.save()
            return JsonResponse({'success': True})
        except StaffNotification.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Notification not found'})
    return JsonResponse({'error': 'Invalid method'}, status=400)


@login_required
@staff_required
def mark_all_notifications_read(request):
    """Mark all notifications as read"""
    if request.method == 'POST':
        StaffNotification.objects.filter(staff=request.user.staff_profile, is_read=False).update(is_read=True,
                                                                                                 read_at=timezone.now())
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'Invalid method'}, status=400)


# ==================== PROFILE VIEWS ====================

@login_required
@staff_required
def staff_profile_view(request):
    """Staff profile page"""
    staff_profile_obj = request.user.staff_profile

    applications_reviewed = StaffActivityLog.objects.filter(staff=staff_profile_obj, entity_type='application',
                                                            action='REVIEW').count()
    loans_processed = StaffActivityLog.objects.filter(staff=staff_profile_obj, entity_type='loan',
                                                      action='CREATE').count()
    payments_processed = StaffActivityLog.objects.filter(staff=staff_profile_obj, entity_type='payment',
                                                         action='CREATE').count()
    recent_activities = StaffActivityLog.objects.filter(staff=staff_profile_obj).order_by('-created_at')[:10]

    context = {
        'staff_profile': staff_profile_obj,
        'applications_reviewed': applications_reviewed,
        'loans_processed': loans_processed,
        'payments_processed': payments_processed,
        'recent_activities': recent_activities,
    }

    return render(request, 'staff/profile/index.html', context)


@login_required
@staff_required
def edit_profile(request):
    """Edit profile page"""
    staff_profile = request.user.staff_profile

    context = {
        'staff_profile': staff_profile,
    }

    return render(request, 'staff/profile/edit.html', context)


@login_required
@staff_required
def update_profile(request):
    """Update staff profile"""
    if request.method == 'POST':
        staff_profile_obj = request.user.staff_profile
        user = request.user

        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.save()

        staff_profile_obj.contact_number = request.POST.get('contact_number', staff_profile_obj.contact_number)
        staff_profile_obj.save()

        log_activity(staff_profile_obj, 'UPDATE', 'profile', staff_profile_obj.id, request)

        messages.success(request, 'Profile updated successfully!')
        return redirect('staff:staff_profile')

    return redirect('staff:staff_profile')


@login_required
@staff_required
def upload_avatar(request):
    """Upload profile picture"""
    if request.method == 'POST' and request.FILES.get('avatar'):
        staff_profile = request.user.staff_profile
        if staff_profile.profile_picture:
            import os
            if os.path.isfile(staff_profile.profile_picture.path):
                os.remove(staff_profile.profile_picture.path)
        staff_profile.profile_picture = request.FILES['avatar']
        staff_profile.save()
        messages.success(request, 'Profile picture updated successfully!')
    return redirect('staff:staff_profile')


@login_required
@staff_required
def change_password(request):
    """Change staff password"""
    if request.method == 'POST':
        user = request.user
        current = request.POST.get('current_password')
        new = request.POST.get('new_password')
        confirm = request.POST.get('confirm_password')

        if not user.check_password(current):
            messages.error(request, 'Current password is incorrect.')
        elif new != confirm:
            messages.error(request, 'New passwords do not match.')
        elif len(new) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
        else:
            user.set_password(new)
            user.save()
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, user)
            messages.success(request, 'Password changed successfully!')

        return redirect('staff:staff_profile')

    return redirect('staff:staff_profile')


@login_required
@staff_required
def logout_all_devices(request):
    """Logout from all devices"""
    if request.method == 'POST':
        from django.contrib.auth import logout
        from django.contrib.sessions.models import Session

        user = request.user
        sessions = Session.objects.filter(
            session_key__in=[s.session_key for s in Session.objects.all()
                             if s.get_decoded().get('_auth_user_id') == str(user.id)]
        )
        sessions.delete()
        logout(request)

        messages.success(request, 'Logged out from all devices.')
        return redirect('main:login')

    return redirect('staff:staff_profile')


# ==================== 2FA AUTHENTICATOR VIEWS ====================

@login_required
@staff_required
def setup_2fa(request):
    """Setup Two-Factor Authentication for staff"""
    staff_profile = request.user.staff_profile

    if request.method == 'POST':
        device = TOTPDevice.objects.filter(user=request.user).first()
        if not device:
            device = TOTPDevice.objects.create(
                user=request.user,
                name='Default',
                confirmed=False
            )

        otp_code = request.POST.get('otp_code', '')
        if device.verify_token(otp_code):
            device.confirmed = True
            device.save()

            staff_profile.otp_enabled = True
            staff_profile.otp_enabled_at = timezone.now()
            staff_profile.otp_secret = device.bin_key.hex()

            backup_codes = []
            for i in range(10):
                code = secrets.token_hex(4).upper()
                formatted_code = f"{code[:5]}-{code[5:]}"
                backup_codes.append(formatted_code)
            staff_profile.otp_backup_codes = backup_codes
            staff_profile.save()

            messages.success(request, 'Two-Factor Authentication enabled successfully!')
            return redirect('staff:staff_profile')
        else:
            messages.error(request, 'Invalid verification code. Please try again.')
            return redirect('staff:setup_2fa')

    device = TOTPDevice.objects.filter(user=request.user).first()
    if not device:
        device = TOTPDevice.objects.create(
            user=request.user,
            name='Default',
            confirmed=False
        )

    provisioning_uri = device.config_url

    qr = qrcode.QRCode(box_size=10, border=4)
    qr.add_data(provisioning_uri)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        'staff_profile': staff_profile,
        'qr_code': qr_base64,
        'secret_key': device.bin_key.hex(),
        'provisioning_uri': provisioning_uri,
    }

    return render(request, 'staff/profile/setup_2fa.html', context)


@login_required
@staff_required
def disable_2fa(request):
    """Disable Two-Factor Authentication"""
    if request.method == 'POST':
        staff_profile = request.user.staff_profile
        TOTPDevice.objects.filter(user=request.user).delete()

        staff_profile.otp_enabled = False
        staff_profile.otp_secret = None
        staff_profile.otp_backup_codes = []
        staff_profile.save()

        messages.success(request, 'Two-Factor Authentication has been disabled.')
        return redirect('staff:staff_profile')

    return redirect('staff:staff_profile')


@login_required
@staff_required
def generate_backup_codes(request):
    """Generate new backup codes for 2FA"""
    if request.method == 'POST':
        staff_profile = request.user.staff_profile

        if not staff_profile.otp_enabled:
            messages.error(request, '2FA is not enabled. Please enable it first.')
            return redirect('staff:setup_2fa')

        new_backup_codes = []
        for i in range(10):
            code = secrets.token_hex(4).upper()
            formatted_code = f"{code[:5]}-{code[5:]}"
            new_backup_codes.append(formatted_code)

        staff_profile.otp_backup_codes = new_backup_codes
        staff_profile.save()

        messages.success(request, 'New backup codes generated!')

    return redirect('staff:staff_profile')


@login_required
@staff_required
def verify_2fa_login(request):
    """Verify 2FA code during login"""
    if request.method == 'POST':
        otp_code = request.POST.get('otp_code', '')
        user = request.user

        device = TOTPDevice.objects.filter(user=user, confirmed=True).first()
        if device and device.verify_token(otp_code):
            request.session['2fa_verified'] = True
            return redirect(request.GET.get('next', '/staff/'))

        staff_profile = user.staff_profile
        if otp_code in staff_profile.otp_backup_codes:
            codes = staff_profile.otp_backup_codes
            codes.remove(otp_code)
            staff_profile.otp_backup_codes = codes
            staff_profile.save()
            request.session['2fa_verified'] = True
            return redirect(request.GET.get('next', '/staff/'))

        messages.error(request, 'Invalid 2FA code. Please try again.')

    return render(request, 'staff/profile/verify_2fa.html')


# ==================== API ENDPOINTS ====================

@login_required
@staff_required
def member_search(request):
    """Search members"""
    q = request.GET.get('q', '')
    members = Member.objects.filter(
        Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(membership_number__icontains=q)
    )[:10]

    results = [{'id': m.id, 'name': f"{m.last_name}, {m.first_name}", 'membership_id': m.membership_number} for m in
               members]

    return JsonResponse({'members': results})


@login_required
@staff_required
def loan_status_api(request, loan_id):
    """Get loan status"""
    loan = get_object_or_404(Loan, pk=loan_id)

    return JsonResponse({
        'id': loan.id,
        'number': loan.loan_number,
        'remaining_balance': float(loan.remaining_balance),
        'next_due_date': loan.due_date.strftime('%Y-%m-%d') if loan.due_date else None,
        'monthly_payment': float(loan.monthly_payment),
        'status': loan.status,
    })


@login_required
@staff_required
def calculate_payment_breakdown(request):
    """Calculate payment breakdown"""
    amount = float(request.GET.get('amount', 0))
    balance = float(request.GET.get('balance', 0))

    interest = balance * 0.0166667
    principal = amount - interest if amount > interest else 0
    remaining = balance - principal if amount > interest else balance

    return JsonResponse({
        'penalty': 0,
        'interest': round(interest, 2),
        'principal': round(principal, 2),
        'remaining_balance': round(remaining, 2)
    })


@login_required
@staff_required
def application_api(request, pk):
    """API endpoint for application details"""
    try:
        app = get_object_or_404(LoanApplication, pk=pk)
        return JsonResponse({
            'id': app.id,
            'application_id': app.application_id,
            'member_name': f"{app.member.last_name}, {app.member.first_name}",
            'member_id': app.member.membership_number,
            'loan_product': app.loan_product.name if app.loan_product else 'N/A',
            'amount': float(app.amount),
            'approved_line': float(app.approved_line) if app.approved_line else 0,
            'status': app.status,
            'applied_date': app.applied_date.strftime('%Y-%m-%d'),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@staff_required
def bulk_forward(request):
    """Bulk forward applications to committee"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            app_ids = data.get('application_ids', [])
            updated = LoanApplication.objects.filter(id__in=app_ids, status='pending_staff_review').update(
                status='with_committee')
            return JsonResponse({'success': True, 'count': updated})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'error': 'Invalid method'}, status=400)


@login_required
@staff_required
def bulk_reject(request):
    """Bulk reject applications"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            app_ids = data.get('application_ids', [])
            reason = data.get('reason', 'Bulk rejection')
            updated = LoanApplication.objects.filter(id__in=app_ids).update(status='rejected', staff_remarks=reason)
            return JsonResponse({'success': True, 'count': updated})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'error': 'Invalid method'}, status=400)


@login_required
@staff_required
def validate_co_maker(request):
    """Validate co-maker by ID"""
    co_maker_id = request.GET.get('co_maker_id', '')

    try:
        co_maker = Member.objects.get(membership_number=co_maker_id, is_active=True)
        return JsonResponse({
            'valid': True,
            'name': f"{co_maker.last_name}, {co_maker.first_name}",
            'membership_id': co_maker.membership_number,
            'message': 'Co-maker validated successfully'
        })
    except Member.DoesNotExist:
        return JsonResponse({
            'valid': False,
            'message': 'Co-maker not found or inactive'
        })


@login_required
def staff_logout(request):
    """Staff logout"""
    from django.contrib.auth import logout
    logout(request)
    return redirect('main:landing')